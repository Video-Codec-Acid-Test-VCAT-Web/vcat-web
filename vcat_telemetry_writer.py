import atexit
import json
import os
import re
from datetime import datetime
from enum import Enum
from typing import Any, cast, Dict, List

from openpyxl import __version__, Workbook
from openpyxl.worksheet.worksheet import Worksheet


print(f"[DEBUG] openpyxl version: {__version__}")
print(f"[DEBUG] Workbook type: {type(Workbook)}")

from vcat_logging import logger  # ✅ shared logger
from vcat_telemetry_data_models import *

__all__ = [
    "append_telemetry",
    "close_telemetry_excel",
    "create_telemetry_excel",
    "TelemetrySheet",
    "export_telemetry",
]


# Sheet name enum for safe usage
class TelemetrySheet(str, Enum):
    SUMMARY = "Summary"
    THERMAL = "Thermal"
    BATTERY = "Battery"
    CPU_USAGE = "CPU Usage"
    CPU_FREQ = "CPU Frequency"
    FRAME_DROPS = "Frame Drops"
    MEMORY = "Memory"


# Maintain open workbooks in memory per device
workbook_handles: Dict[str, Workbook] = {}
file_paths: Dict[str, str] = {}


def format_device_filename(device_id: str, manufacturer: str, model: str) -> str:
    def sanitize(s: str) -> str:
        return re.sub(r"[^A-Za-z0-9._-]", "_", s.strip())

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return (
        f"{sanitize(device_id)}_{sanitize(manufacturer)}_{sanitize(model)}_{timestamp}"
    )


def export_telemetry(
    session_id, device_id, telemetry_data: TelemetryData, output_path: str
) -> str:

    create_telemetry_excel_at_path(telemetry_data, output_path)

    append_telemetry(
        session_id,
        TelemetrySheet.THERMAL,
        [[entry.elapsed_time, entry.status] for entry in telemetry_data.system_thermal_status],
    )

    append_telemetry(
        session_id,
        TelemetrySheet.BATTERY,
        [[entry.elapsed_time, entry.level, entry.charge_count, entry.current_ma, entry.battery_temp] for entry in telemetry_data.battery_data],
    )

    if telemetry_data.system_memory and telemetry_data.app_memory:
        rows = []

        for sys_entry, app_entry in zip(
            telemetry_data.system_memory, telemetry_data.app_memory
        ):
            row = [
                sys_entry.elapsed_time,
                sys_entry.used_kb,
                app_entry.used_kb,
            ]
            rows.append(row)

        append_telemetry(telemetry_data.owner_session_id, TelemetrySheet.MEMORY, rows)

    if telemetry_data.cpu_usage:
        rows = []
        for entry in telemetry_data.cpu_usage:
            row = [entry.elapsed_time]

            # Add total CPU usage first
            row.append(entry.usage_pct.get("cpu", 0.0))

            # Add per-core usage in consistent order
            core_keys = sorted(
                k for k in entry.usage_pct if k.startswith("cpu") and k != "cpu"
            )
            row.extend([entry.usage_pct.get(k, 0.0) for k in core_keys])

            rows.append(row)

        append_telemetry(
            telemetry_data.owner_session_id, TelemetrySheet.CPU_USAGE, rows
        )

    if telemetry_data.cpu_freq:
        rows = []
        for entry in telemetry_data.cpu_freq:
            row = [entry.elapsed_time]
            core_keys = sorted(entry.frequencies.keys())
            row.extend([entry.frequencies[k] for k in core_keys])
            rows.append(row)

        append_telemetry(telemetry_data.owner_session_id, TelemetrySheet.CPU_FREQ, rows)

    if telemetry_data.frame_drops:
        append_telemetry(
            telemetry_data.owner_session_id,
            TelemetrySheet.FRAME_DROPS,
            [
                [entry.elapsed_time, entry.delta_framedrops]
                for entry in telemetry_data.frame_drops
            ],
        )

    close_telemetry_excel(session_id)

    return output_path


def create_telemetry_excel(telemetry_data: TelemetryData) -> str:
    base_filename = format_device_filename(
        telemetry_data.device_id,
        telemetry_data.device_info.manufacturer,
        telemetry_data.device_info.model,
    )

    os.makedirs("output", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"output/{base_filename}_{timestamp}.xlsx"

    return create_telemetry_excel_at_path(telemetry_data, filename)


def create_telemetry_excel_at_path(
    telemetry_data: TelemetryData, file_path_name: str
) -> str:

    wb = Workbook()

    print(f"[DEBUG] Created workbook: {wb}")
    print(f"[DEBUG] Sheet names: {wb.sheetnames}")
    print(f"[DEBUG] Active sheet: {wb.active}")

    sheet = cast(Worksheet, wb.active)

    if sheet is None:
        raise RuntimeError("Workbook has no active worksheet!")

    sheet.title = TelemetrySheet.SUMMARY.value

    # Write device info into the SUMMARY sheet
    device_info_dict = telemetry_data.device_info.to_dict()
    session_info_dict = telemetry_data.session_info.to_dict()
    test_conditions_dict = telemetry_data.test_conditions.to_dict()

    header_dict = {
        "header_version": telemetry_data.version,
        "device_info": device_info_dict,
        "session_info": session_info_dict,
        "test_conditions": test_conditions_dict,
    }

    summary_sheet = sheet

    device_info_json = json.dumps(header_dict, indent=1, sort_keys=False)

    for line in device_info_json.splitlines():
        summary_sheet.append([line])

    summary_sheet.append([])

    for sheet_name in [
        TelemetrySheet.THERMAL.value,
        TelemetrySheet.BATTERY.value,
        TelemetrySheet.CPU_USAGE.value,
        TelemetrySheet.CPU_FREQ.value,
        TelemetrySheet.FRAME_DROPS.value,
        TelemetrySheet.MEMORY.value,
    ]:
        wb.create_sheet(sheet_name)

    ccore_labels = [
        f"cpu{core.core_id}" for core in telemetry_data.device_info.cpu.cores
    ]

    wb[TelemetrySheet.THERMAL.value].append(
        ["Elapsed Time (s)", "Status"]
    )

    wb[TelemetrySheet.BATTERY.value].append(["Elapsed Time (s)", "battery.level (%)", "battery.charge_counter", "battery.milliamps", "battery.temperature"])
    wb[TelemetrySheet.CPU_USAGE.value].append(
        ["Elapsed Time (s)", "total"] + ccore_labels
    )
    wb[TelemetrySheet.CPU_FREQ.value].append(["Elapsed Time (s)"] + ccore_labels)
    wb[TelemetrySheet.FRAME_DROPS.value].append(
        ["Elapsed Time (s)", "Delta Frame Drops"]
    )
    wb[TelemetrySheet.MEMORY.value].append(
        ["Elapsed Time (s)", "Total KB", "Used KB", "App KB"]
    )

    wb.save(file_path_name)
    workbook_handles[telemetry_data.owner_session_id] = wb
    file_paths[telemetry_data.owner_session_id] = file_path_name
    logger.info(f"📁 Created telemetry file: {file_path_name}")
    return file_path_name


def append_telemetry(session_id: str, sheet: TelemetrySheet, rows: List[List[Any]]):
    if session_id not in workbook_handles:
        logger.critical(f"Device workbook not initialized: {session_id}")

    wb = workbook_handles[session_id]
    sheet_name = sheet.value

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet does not exist: {sheet_name}")

    ws = wb[sheet_name]
    for row in rows:
        ws.append(row)

    wb.save(file_paths[session_id])


def close_telemetry_excel(session_id: str):
    if session_id in workbook_handles:
        wb = workbook_handles.pop(session_id)
        path = file_paths.pop(session_id, None)
        if path:
            wb.save(path)
            logger.info(f"✅ Closed workbook for session: {session_id}")


def cleanup_all_workbooks():
    for session_id in list(workbook_handles.keys()):
        close_telemetry_excel(session_id)
    logger.info("🧹 All telemetry workbooks have been closed.")


atexit.register(cleanup_all_workbooks)
